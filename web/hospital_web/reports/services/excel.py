from __future__ import annotations

import io
from datetime import datetime
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from scheduling.models import Appointment


def _setup_styles():
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(name="Arial", size=10, bold=True)
    subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    data_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    return header_font, header_fill, subheader_font, subheader_fill, data_font, center, left, border


def export_doctor_month(*, doctor_name: str, specialization: str, year: int, month: int, appointments: Iterable[Appointment]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Записи_{month:02d}_{year}"

    header_font, header_fill, subheader_font, subheader_fill, data_font, center, left, border = _setup_styles()

    ws.merge_cells("A1:H1")
    ws["A1"] = f"ЗАПИСИ ВРАЧА ЗА {month:02d}.{year}"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = center

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Врач: {doctor_name} | Специализация: {specialization}"
    ws["A2"].font = subheader_font
    ws["A2"].fill = subheader_fill
    ws["A2"].alignment = center

    headers = ["№", "Дата", "Время", "Пациент", "Телефон", "Кабинет", "Длительность (мин)", "Заметки"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center
        cell.border = border

    row = 5
    i = 1
    for a in appointments:
        ws.cell(row=row, column=1, value=i).border = border
        ws.cell(row=row, column=2, value=a.appointment_date.strftime("%d.%m.%Y")).border = border
        ws.cell(row=row, column=3, value=a.appointment_time.strftime("%H:%M")).border = border
        ws.cell(row=row, column=4, value=a.patient.name).border = border
        ws.cell(row=row, column=5, value=a.patient.phone).border = border
        ws.cell(row=row, column=6, value=a.room.name).border = border
        ws.cell(row=row, column=7, value=a.duration_minutes).border = border
        ws.cell(row=row, column=8, value=a.notes or "").border = border
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = left if col in [4, 5, 8] else center
        row += 1
        i += 1

    # Serialize to bytes
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_all_doctors_month(*, year: int, month: int, appointments: Iterable[Appointment]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Все_записи_{month:02d}_{year}"

    header_font, header_fill, subheader_font, subheader_fill, data_font, center, left, border = _setup_styles()

    ws.merge_cells("A1:J1")
    ws["A1"] = f"ЗАПИСИ ВСЕХ ВРАЧЕЙ ЗА {month:02d}.{year}"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = center

    headers = ["№", "Дата", "Время", "Врач", "Специализация", "Пациент", "Телефон", "Кабинет", "Длительность (мин)", "Заметки"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center
        cell.border = border

    row = 4
    i = 1
    for a in appointments:
        ws.cell(row=row, column=1, value=i).border = border
        ws.cell(row=row, column=2, value=a.appointment_date.strftime("%d.%m.%Y")).border = border
        ws.cell(row=row, column=3, value=a.appointment_time.strftime("%H:%M")).border = border
        ws.cell(row=row, column=4, value=a.doctor.name).border = border
        ws.cell(row=row, column=5, value=a.doctor.specialization).border = border
        ws.cell(row=row, column=6, value=a.patient.name).border = border
        ws.cell(row=row, column=7, value=a.patient.phone).border = border
        ws.cell(row=row, column=8, value=a.room.name).border = border
        ws.cell(row=row, column=9, value=a.duration_minutes).border = border
        ws.cell(row=row, column=10, value=a.notes or "").border = border
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = left if col in [4, 5, 6, 7, 10] else center
        row += 1
        i += 1

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

